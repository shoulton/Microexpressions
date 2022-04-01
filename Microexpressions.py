import torch
import PIL as pil
import openpyxl as xl
import torch.transforms as transforms

def makeData():
    wb = xl.load_workbook(filename='CASME/CASME2-coding-20190701.xlsx')
    sheet = wb.active
    max_row = sheet.max_row

    pictures = []
    emotions = []

    for i in range(1, max_row + 1):
        subject = sheet.cell(row=i, column=1).value
        filename = sheet.cell(row = i, column=2).value
        apex = sheet.cell(row = i, column=5).value
        emotion = sheet.cell(row = 1, column=9).value
        path = makePath(subject, filename, apex).value
        pil_image = pil.Image.open(path)
        rgb_image = transforms.toTensor(pil_image)
        pictures.append(rgb_image)

def makePath(subject, filename, apexFrame):
    path = ""
    if subject == 1:
        path = f"CASME/Cropped-updated/Cropped/sub01/{filename}/reg_img{apexFrame}.jpg"
    if subject == 2:
        path = f"CASME/Cropped-updated/Cropped/sub02/{filename}/reg_img{apexFrame}.jpg"
    if subject == 3:
        path = f"CASME/Cropped-updated/Cropped/sub03/{filename}/reg_img{apexFrame}.jpg"
    if subject == 4:
        path = f"CASME/Cropped-updated/Cropped/sub04/{filename}/reg_img{apexFrame}.jpg"
    if subject == 5:
        path = f"CASME/Cropped-updated/Cropped/sub05/{filename}/reg_img{apexFrame}.jpg"
    if subject == 6:
        path = f"CASME/Cropped-updated/Cropped/sub06/{filename}/reg_img{apexFrame}.jpg"
    if subject == 7:
        path = f"CASME/Cropped-updated/Cropped/sub07/{filename}/reg_img{apexFrame}.jpg"
    if subject == 8:
        path = f"CASME/Cropped-updated/Cropped/sub08/{filename}/reg_img{apexFrame}.jpg"
    if subject == 9:
        path = f"CASME/Cropped-updated/Cropped/sub09/{filename}/reg_img{apexFrame}.jpg"
    if subject == 10:
        path = f"CASME/Cropped-updated/Cropped/sub10/{filename}/reg_img{apexFrame}.jpg"
    if subject == 11:
        path = f"CASME/Cropped-updated/Cropped/sub11/{filename}/reg_img{apexFrame}.jpg"
    if subject == 12:
        path = f"CASME/Cropped-updated/Cropped/sub12/{filename}/reg_img{apexFrame}.jpg"
    if subject == 13:
        path = f"CASME/Cropped-updated/Cropped/sub13/{filename}/reg_img{apexFrame}.jpg"
    if subject == 14:
        path = f"CASME/Cropped-updated/Cropped/sub14/{filename}/reg_img{apexFrame}.jpg"
    if subject == 15:
        path = f"CASME/Cropped-updated/Cropped/sub15/{filename}/reg_img{apexFrame}.jpg"
    if subject == 16:
        path = f"CASME/Cropped-updated/Cropped/sub16/{filename}/reg_img{apexFrame}.jpg"
    if subject == 17:
        path = f"CASME/Cropped-updated/Cropped/sub17/{filename}/reg_img{apexFrame}.jpg"
    if subject == 18:
        path = f"CASME/Cropped-updated/Cropped/sub18/{filename}/reg_img{apexFrame}.jpg"
    if subject == 19:
        path = f"CASME/Cropped-updated/Cropped/sub19/{filename}/reg_img{apexFrame}.jpg"
    if subject == 20:
        path = f"CASME/Cropped-updated/Cropped/sub20/{filename}/reg_img{apexFrame}.jpg"
    if subject == 21:
        path = f"CASME/Cropped-updated/Cropped/sub21/{filename}/reg_img{apexFrame}.jpg"
    if subject == 22:
        path = f"CASME/Cropped-updated/Cropped/sub22/{filename}/reg_img{apexFrame}.jpg"
    if subject == 23:
        path = f"CASME/Cropped-updated/Cropped/sub23/{filename}/reg_img{apexFrame}.jpg"
    if subject == 24:
        path = f"CASME/Cropped-updated/Cropped/sub24/{filename}/reg_img{apexFrame}.jpg"
    if subject == 25:
        path = f"CASME/Cropped-updated/Cropped/sub25/{filename}/reg_img{apexFrame}.jpg"
    if subject == 26:
        path = f"CASME/Cropped-updated/Cropped/sub26/{filename}/reg_img{apexFrame}.jpg"

    return path


if __name__ == "__main__":
    makeData()